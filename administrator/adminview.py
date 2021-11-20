
from django.contrib import messages
from django.core.files.storage import FileSystemStorage
from django.http import HttpResponse, HttpResponseRedirect
from django.shortcuts import render
from django.urls import reverse
from administrator.models import CustomUser,Admin
from classes.models import Courses,Subjects,SessionYearModel
from staff.models import Staffs,LeaveReportStaff
from student.models import Students,Students_bulk,LeaveReportStudent
from student.form import AddStudentForm, EditStudentForm
from django.dispatch import receiver
import openpyxl
from openpyxl import load_workbook
from io import BytesIO

# def admin_home(request):
#     return render(request,"hod_template/base_template.html")
def admin_home(request):
    student_count1=Students.objects.all().count()
    staff_count=Staffs.objects.all().count()
    subject_count=Subjects.objects.all().count()
    course_count=Courses.objects.all().count()
    course_all=Courses.objects.all()
    course_name_list=[]
    subject_count_list=[]
    student_count_list_in_course=[]
    for course in course_all:
        subjects=Subjects.objects.filter(course_id=course.id).count()
        students=Students.objects.filter(course_id=course.id).count()
        course_name_list.append(course.course_name)
        subject_count_list.append(subjects)
        student_count_list_in_course.append(students)
    subjects_all=Subjects.objects.all()
    subject_list=[]
    student_count_list_in_subject=[]
    for subject in subjects_all:
        course = Courses.objects.get(id=subject.course_id.id)
        student_count=Students.objects.filter(course_id=course.id).count()
        subject_list.append(subject.subject_name)
        student_count_list_in_subject.append(student_count)

    # staffs=Staffs.objects.all()
    # attendance_present_list_staff=[]
    # attendance_absent_list_staff=[]
    # staff_name_list=[]
    # for staff in staffs:
    #     subject_ids=Subjects.objects.filter(staff_id=staff.admin.id)
    #     # attendance=Attendance.objects.filter(subject_id__in=subject_ids).count()
    #     leaves=LeaveReportStaff.objects.filter(staff_id=staff.id,leave_status=1).count()
    #     attendance_present_list_staff.append(attendance)
    #     attendance_absent_list_staff.append(leaves)
    #     staff_name_list.append(staff.admin.username)

    # students_all=Students.objects.all()
    # attendance_present_list_student=[]
    # attendance_absent_list_student=[]
    # student_name_list=[]
    # for student in students_all:
    #     # attendance=AttendanceReport.objects.filter(student_id=student.id,status=True).count()
    #     # absent=AttendanceReport.objects.filter(student_id=student.id,status=False).count()
    #     leaves=LeaveReportStudent.objects.filter(student_id=student.id,leave_status=1).count()
    #     attendance_present_list_student.append(attendance)
    #     attendance_absent_list_student.append(leaves+absent)
    #     student_name_list.append(student.admin.username)
    return render(request,"hod_template/home_content.html",{"student_count":student_count1,"staff_count":staff_count,"subject_count":subject_count,"course_count":course_count,"course_name_list":course_name_list,"subject_count_list":subject_count_list,"student_count_list_in_course":student_count_list_in_course,"student_count_list_in_subject":student_count_list_in_subject})



def add_student(request):
    form=AddStudentForm()
    return render(request,"hod_template/add_student_template.html",{"form":form})


def add_bulk_student(request):
    return render(request,"hod_template/add_bulk_students.html")

def add_bulk_student_save(request):
    if request.method == 'POST':
        wb = load_workbook(filename=request.FILES['file'].file)
        sheet_obj=wb.active
        m_row = sheet_obj.max_row 
        for i in range(2, m_row + 1):
            reg_no = sheet_obj.cell(row = i, column = 1)
            bate= sheet_obj.cell(row = i, column = 2)
            first_name = sheet_obj.cell(row = i, column = 3)
            gender=sheet_obj.cell(row = i, column = 4)
            section=sheet_obj.cell(row = i, column = 5)
            obj=Courses.objects.filter(course_name=section)
            if obj.count()>0:
                for i in obj.count():
                    request.session["id"]=obj[i].id
            check=CustomUser.objects.filter(username=reg_no.value).exists()
            if not check:
                user=CustomUser.objects.create_user(username=reg_no.value,password=bate.value,first_name=first_name.value,user_type=3)
                user.students.bdate=bate.value
                user.students.gender=gender.value
                user.students.course_id=id
                user.save()
        messages.success(request,"Successfully Added Student")
        students = Students.objects.all()
        courses = Courses.objects.all()
        return render(request, "hod_template/manage_student_template.html", {"students": students, "courses": courses})
    students = Students.objects.all()
    courses = Courses.objects.all()
    messages.success(request, "Students already exists")
    return render(request, "hod_template/manage_student_template.html", {"students": students, "courses": courses})

def add_staff_save(request):
    if request.method!="POST":
        return HttpResponse("Method Not Allowed")
    else:
        first_name=request.POST.get("first_name")
        username=request.POST.get("username")
        email=request.POST.get("email")
        password=request.POST.get("sbdate")
        sbdate=request.POST.get("sbdate")
        address=request.POST.get("address")
        mob=request.POST.get('mob')
        adhar = request.POST.get('adhar')
        gd = request.POST.get('gd')
        bd = request.POST.get('bd')
        mt = request.POST.get('mt')
        ms = request.POST.get('ms')
        re = request.POST.get('re')
        caste = request.POST.get('caste')
        cnt = request.POST.get('cnt')
        st =  request.POST.get('st')
        district = request.POST.get('district')
        Qualification = request.POST.get('Qualification')
        profile_pic = request.FILES['profile_pic']
        fs=FileSystemStorage()
        filename=fs.save(profile_pic.name,profile_pic)
        profile_pic_url=fs.url(filename)
        try:
            user=CustomUser.objects.create_user(username=username,password=password,first_name=first_name,user_type=2)
            user.staffs.address=address
            user.staffs.email=email
            user.staffs.sbdate=sbdate
            user.staffs.mobileno=mob
            user.staffs.adhar_no=adhar
            user.staffs.gender=gd
            user.staffs.blood_group=bd
            user.staffs.mother_tounge=mt
            user.staffs.marital_status=ms
            user.staffs.relegion=re
            user.staffs.caste=caste
            user.staffs.country=cnt
            user.staffs.state=st
            user.staffs.district=district
            user.staffs.qualification=Qualification
            user.staffs.profile_pic=profile_pic_url
            user.save()
            messages.success(request,"Successfully Added Staff")
            return HttpResponseRedirect(reverse("manage_staff"))
        except:
            messages.error(request,"Failed to Add Staff")
            return HttpResponseRedirect(reverse("manage_staff"))

def add_student_save(request):
    if request.method!="POST":
        return HttpResponse("Method Not Allowed")
    else:
        first_name = request.POST.get("first_name")
        username = request.POST.get("username")
        email = request.POST.get("email")
        password = request.POST.get("bdate")
        bdate = request.POST.get("bdate")
        address = request.POST.get("address")
        mob = request.POST.get('mob')
        adhar = request.POST.get('adhar')
        gd = request.POST.get('gd')
        bd = request.POST.get('bd')
        mt = request.POST.get('mt')
        ms = request.POST.get('ms')
        re = request.POST.get('re')
        course_id= request.POST.get('cour_name')
        caste = request.POST.get('caste')
        cnt = request.POST.get('cnt')
        st = request.POST.get('st')
        father= request.POST.get('father')
        mother = request.POST.get('mother')
        fatherno = request.POST.get('fatherno')
        motherno = request.POST.get('motherno')
        fatherocc = request.POST.get('fatherocc')
        Motherocc = request.POST.get('Motherocc')
        income= request.POST.get('income')
        district = request.POST.get('district')
        '''
        profile_pic = request.POST.get('profile_pic')
        fs = FileSystemStorage()
        filename = fs.save(profile_pic.name, profile_pic)
        profile_pic_url = fs.url(filename)
        '''
        try:
            user = CustomUser.objects.create_user(username=username, password=password, first_name=first_name,user_type=3)
            user.students.address = address
            user.students.email = email
            user.students.bdate = bdate
            user.students.mobileno = mob
            user.students.adhar_no = adhar
            user.students.gender = gd
            user.students.blood_group = bd
            user.students.mother_tounge = mt
            user.students.marital_status = ms
            user.students.relegion = re
            user.students.caste = caste
            user.students.country = cnt
            user.students.state = st
            user.students.fathername=father
            user.students.mothername = mother
            user.students.parents_mobileno1=fatherno
            user.students.parents_mobileno2=motherno
            user.students.fatheroccupation=fatherocc
            user.students.motheroccupation=Motherocc
            user.students.fatherincome=income
            user.students.district = district
            course_obj = Courses.objects.get(id=course_id)
            user.students.course_id = course_obj
            user.students.gender = gd
            #user.students.profile_pic = profile_pic_url
            user.save()
            messages.success(request, "Successfully Added Student")
            return HttpResponseRedirect(reverse("manage_student"))
        except:
            messages.error(request, "Failed to Add Student")
            return HttpResponseRedirect(reverse("manage_student"))




def manage_staff(request):
    staffs=Staffs.objects.all()
    return render(request,"hod_template/manage_staff_template.html",{"staffs":staffs})

def manage_student(request):
    students=Students.objects.all()
    courses=Courses.objects.all()
    return render(request,"hod_template/manage_student_template.html",{"students":students,"courses":courses})

def manage_student_filter(request):
    first_name=request.GET.get("first_name")
    username=request.GET.get("username")
    course_name=request.GET.get("course_name")
    data=Students.objects.all()
    courses=Courses.objects.all()

    if(first_name !=''):
        data=data.filter(admin__first_name__icontains=first_name)
    if(username !=''):
        data=data.filter(admin__username__icontains=username)  
    if(course_name !=''):
        data=data.filter(course_id__course_name__icontains=course_name)
   
    
    return render(request,"hod_template/manage_student_template.html",{"students":data,"courses":courses})

def manage_course(request):
    staffs = CustomUser.objects.filter(user_type=2)
    courses=Courses.objects.all()
    return render(request,"hod_template/manage_course_template.html",{"courses":courses,"staffs":staffs})

def manage_subject(request):
    subjects=Subjects.objects.all()
    return render(request,"hod_template/manage_subject_template.html",{"subjects":subjects})

def manage_class_subject_staff(request):
    subjects = Subjects.objects.all()
    courses = Courses.objects.all()
    staffs=Staffs.objects.all()
    return render(request,"hod_template/add_classvise_subject_staff.html",{"subjects":subjects,"courses":courses,"staffs":staffs})

def edit_staff(request,staff_id):
    staff=Staffs.objects.get(admin=staff_id)
    return render(request,"hod_template/edit_staff_template.html",{"staff":staff,"id":staff_id})

def delete_staff(request,staff_id):
    CustomUser.objects.filter(id=staff_id).delete()
    Staffs.objects.filter(admin=staff_id).delete()
    staff = Staffs.objects.all()
    return render(request,"hod_template/manage_staff_template.html",{"staffs":staff})
'''
def delete_student_confirm(request):
    return render(request,"student_template/delete_student_confirm.html")
'''
def delete_student(request,student_id):
    Students.objects.filter(admin=student_id).delete()
    CustomUser.objects.filter(id=student_id).delete()
    students = Students.objects.all()
    courses = Courses.objects.all()
    return render(request,"hod_template/manage_student_template.html",{"students":students,"courses":courses})

def delete_class(request,class_id):
    Courses.objects.filter(id=class_id).delete()
    courses = Courses.objects.all()
    return render(request, "hod_template/manage_course_template.html", {"courses": courses})

def delete_subject(request,sub_id):
    Subjects.objects.filter(id=sub_id).delete()
    subjects=Subjects.objects.all()
    return render(request,"hod_template/manage_subject_template.html",{"subjects":subjects})

def edit_staff_save(request):
    if request.method!="POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        staff_id=request.POST.get("staff_id")
        first_name=request.POST.get("first_name")
        last_name=request.POST.get("last_name")
       # email=request.POST.get("email")
        username=request.POST.get("username")
        #address=request.POST.get("address")

        try:
            user=CustomUser.objects.get(id=staff_id)
            user.first_name=first_name
            user.last_name=last_name
            #user.email=email
            user.username=username
            user.save()

            staff_model=Staffs.objects.get(admin=staff_id)
            #staff_model.address=address
            staff_model.save()
            messages.success(request,"Successfully Edited Staff")
            return HttpResponseRedirect(reverse("edit_staff",kwargs={"staff_id":staff_id}))
        except:
            messages.error(request,"Failed to Edit Staff")
            return HttpResponseRedirect(reverse("edit_staff",kwargs={"staff_id":staff_id}))

def edit_student(request,student_id):
    request.session['student_id']=student_id
    student=Students.objects.get(admin=student_id)
    form=EditStudentForm()
    #form.fields['email'].initial=student.admin.email
    form.fields['first_name'].initial=student.admin.first_name
    #form.fields['last_name'].initial=student.admin.last_name
    form.fields['username'].initial=student.admin.username
    #form.fields['address'].initial=student.address
    form.fields['course'].initial=student.course_id.id
    form.fields['sex'].initial=student.gender
    # form.fields['session_start'].initial=student.session_start_year
    # form.fields['session_end'].initial=student.session_end_year
    return render(request,"hod_template/edit_student_template.html",{"form":form,"id":student_id,"username":student.admin.username})

def edit_student_save(request):
    if request.method!="POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        student_id=request.session.get("student_id")
        if student_id==None:
            return HttpResponseRedirect(reverse("manage_student"))

        form=EditStudentForm(request.POST,request.FILES)
        if form.is_valid():
            first_name = form.cleaned_data["first_name"]
            username = form.cleaned_data["username"]                     
            course_id = form.cleaned_data["course"]
            sex = form.cleaned_data["sex"]

            if request.FILES.get('profile_pic',False):
                profile_pic=request.FILES['profile_pic']
                fs=FileSystemStorage()
                filename=fs.save(profile_pic.name,profile_pic)
                profile_pic_url=fs.url(filename)
            else:
                profile_pic_url=None


            try:
                user=CustomUser.objects.get(id=student_id)
                user.first_name=first_name
               # user.last_name=last_name
                user.username=username
                #user.email=email
                user.save()
                student=Students.objects.get(admin=student_id)              
                student.gender=sex
                course=Courses.objects.get(id=course_id)
                student.course_id=course
                if profile_pic_url!=None:
                    student.profile_pic=profile_pic_url
                student.save()
                del request.session['student_id']
                messages.success(request,"Successfully Edited Student")
                return HttpResponseRedirect(reverse("edit_student",kwargs={"student_id":student_id}))
            except:
                messages.error(request,"Failed to Edit Student")
                return HttpResponseRedirect(reverse("edit_student",kwargs={"student_id":student_id}))
        else:
            form=EditStudentForm(request.POST)
            student=Students.objects.get(admin=student_id)
            return render(request,"hod_template/edit_student_template.html",{"form":form,"id":student_id,"username":student.admin.username})




def add_course_save(request):
    if request.method!="POST":
        return HttpResponse("Method Not Allowed")
    else:
        course=request.POST.get("course")
        staff_id=request.POST.get("staff")
        staff=CustomUser.objects.get(id=staff_id)
        try:
            course_model=Courses(course_name=course,staff_id=staff)
            course_model.save()
            messages.success(request,"Successfully Added Course")
            return HttpResponseRedirect(reverse("manage_course"))
        except:
            messages.error(request,"Failed To Add Course")
            return HttpResponseRedirect(reverse("manage_course"))





def add_subject_save(request):
    if request.method!="POST":
        return HttpResponse("<h2>Method Not Allowed</h2>")
    else:
        subject_name=request.POST.get("subject_name")
        '''
        course_id=request.POST.get("course_name")
        course_name=Courses.objects.get(id=course_id)
        staff_id=request.POST.get("staff")
        staff=CustomUser.objects.get(id=staff_id)
        '''
        try:
            subject=Subjects(subject_name=subject_name)
            subject.save()
            messages.success(request,"Successfully Added Subject")
            return HttpResponseRedirect(reverse("manage_subject"))
        except:
            messages.error(request,"Failed to Add Subject")
            return HttpResponseRedirect(reverse("manage_subject"))




def manage_session(request):
    return render(request,"hod_template/manage_session_template.html")
'''
def student_feedback_message(request):
    feedbacks=FeedBackStudent.objects.all()
    return render(request,"hod_template/student_feedback_template.html",{"feedbacks":feedbacks})

def staff_feedback_message(request):
    feedbacks=FeedBackStaffs.objects.all()
    return render(request,"hod_template/staff_feedback_template.html",{"feedbacks":feedbacks})
'''
def staff_leave_view(request):
    leaves=LeaveReportStaff.objects.all()
    return render(request,"hod_template/staff_leave_view.html",{"leaves":leaves})

def student_leave_view(request):
    leaves=LeaveReportStudent.objects.all()
    return render(request,"hod_template/student_leave_view.html",{"leaves":leaves})

def student_approve_leave(request,leave_id):
    leave=LeaveReportStudent.objects.get(id=leave_id)
    leave.leave_status=1
    leave.save()
    return HttpResponseRedirect(reverse("student_leave_view"))

def student_disapprove_leave(request,leave_id):
    leave=LeaveReportStudent.objects.get(id=leave_id)
    leave.leave_status=2
    leave.save()
    return HttpResponseRedirect(reverse("student_leave_view"))


def staff_approve_leave(request,leave_id):
    leave=LeaveReportStaff.objects.get(id=leave_id)
    leave.leave_status=1
    leave.save()
    return HttpResponseRedirect(reverse("staff_leave_view"))

def staff_disapprove_leave(request,leave_id):
    leave=LeaveReportStaff.objects.get(id=leave_id)
    leave.leave_status=2
    leave.save()
    return HttpResponseRedirect(reverse("staff_leave_view"))

def admin_view_attendance(request):
    subjects=Subjects.objects.all()
    session_year_id=SessionYearModel.object.all()
    return render(request,"hod_template/admin_view_attendance.html",{"subjects":subjects,"session_year_id":session_year_id})

def admin_send_notification_student(request):
    students=Students.objects.all()
    return render(request,"hod_template/student_notification.html",{"students":students})

def admin_send_notification_staff(request):
    staffs=Staffs.objects.all()
    return render(request,"hod_template/staff_notification.html",{"staffs":staffs})


from django.db.models.signals import post_save
from django.dispatch import receiver

@receiver(post_save,sender=CustomUser)
def create_user_profile(sender,instance,created,**kwargs):
    if created:
        if instance.user_type==1:
            Admin.objects.create(admin=instance)
        if instance.user_type==2:
            Staffs.objects.create(admin=instance,address="")
        if instance.user_type==3:
            Students.objects.create(admin=instance,course_id=Courses.objects.get(id=1),address="",profile_pic="",gender="")

@receiver(post_save,sender=CustomUser)
def save_user_profile(sender,instance,**kwargs):
    if instance.user_type==1:
        instance.admin.save()
    if instance.user_type==2:
        instance.staffs.save()
    if instance.user_type==3:
        instance.students.save()
